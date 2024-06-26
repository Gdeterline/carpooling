
# Import the required libraries
import pandas as pd
import numpy as np
from pulp import LpProblem, LpVariable, LpBinary, LpInteger, LpMaximize, lpSum
import openpyxl
from openpyxl.styles import PatternFill
import re


red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
dark_grey_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
dark_blue_fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid")



# Load the workbooks and their respective worksheets
input_file = "./Offres_2022.xlsx"
offers_worksheet = pd.read_excel("./Offres_2022.xlsx", sheet_name="Réponses au formulaire 1")
requests_worksheet = pd.read_excel("./Demandes_2022.xlsx", sheet_name="Réponses au formulaire 1")

print("Workbooks loaded successfully")


# Extract the year from the input file name using regular expression
match = re.search(r'(\d{4})', input_file)
if match:
    year = match.group(1)
else:
    raise ValueError("Year not found in the input file name")


# First treatment of the offer sheet data
horaires_offers = offers_worksheet.columns[6:]
#print(horaires_offers)
print()
name_offers = offers_worksheet.loc[:, "Nom"]
print(name_offers)
print()
child_offers = offers_worksheet.loc[:, "Enfant 1":"Enfant 3"]
#print(child_offers)
print()
number_places_offers = offers_worksheet.loc[:, "Places"]
#print(number_places_offers)
print()
availabilities_offers = offers_worksheet.loc[:, "Lundi 8h":"Vendredi 17h30"]
#print(availabilities_offers)
print()



# First treatment of the request sheet data

horaires_requests = requests_worksheet.columns[1:]
print(horaires_requests)
print()
name_requests = requests_worksheet.loc[:, "Prénom"]
print(name_requests)
print()
availabilities_requests = requests_worksheet.loc[:, "Lundi 8h":"Vendredi 18h30"]
print(availabilities_requests)
print()



# Get number of drivers and passengers
nb_requests = 0

number_drivers = len(name_offers)
number_passengers = len(name_requests)
print(f"Number of drivers: {number_drivers}")
print(f"Number of passengers: {number_passengers}")

# number of schedules T
T = requests_worksheet.shape[1] - 1
print("Number of schedules =", T)



# Print total number of requests over 2 weeks
for i in range(number_passengers):
    for j in range(availabilities_requests.shape[1]):
        if availabilities_requests.iloc[i, j] == "Oui":
            nb_requests += 2
        elif availabilities_requests.iloc[i, j] == "Pair":
            nb_requests += 1
        elif availabilities_requests.iloc[i, j] == "Impair":
            nb_requests += 1
            
        
print(f"Number of requests: {nb_requests}")
            

# family matrix with the first column being the passenger number (sheet Demandes_2022) 
# and the second column being the driver number (from the same family) (sheet Offres_2022)

family = np.zeros((number_passengers, 2), dtype=int)

for i in range(number_passengers):
    family[i, 0] = i + 1
    for j in range(number_drivers):
        if name_requests[i] in child_offers.iloc[j].values:
            family[i, 1] = j + 1
            print(f"Passenger {i+1} belongs to the family of driver {j+1}")

print(family)
print()



# places_offered_drivers matrix with the first column being the driver number (sheet Offres_2022) 
# and the second column being the number of places offered by the driver over the two weeks

places_offered_drivers = np.zeros((number_drivers, 2), dtype=int)

for i in range(number_drivers):
    places_offered_drivers[i, 0] = i + 1
    for j in range(availabilities_offers.shape[1]):
        if availabilities_offers.iloc[i, j] == "Oui": 
            places_offered_drivers[i, 1] += 2 * number_places_offers[i]
        elif availabilities_offers.iloc[i, j] == "Impair":
            places_offered_drivers[i, 1] += number_places_offers[i]
        elif availabilities_offers.iloc[i, j] == "Pair":
            places_offered_drivers[i, 1] += number_places_offers[i]

print(places_offered_drivers)
print()




# places_offered_passengers matrix with the first column being the passenger number (sheet Demandes_2022) 
# and the second column being the number of places offered by the driver

places_offered_passengers = np.zeros((number_passengers, 2), dtype=int)

for i in range(number_passengers):
    places_offered_passengers[i, 0] = i + 1
    for j in range(number_drivers):
        if family[i, 1] == places_offered_drivers[j, 0]:
            places_offered_passengers[i, 1] = places_offered_drivers[j, 1]

print(places_offered_passengers)
print()



# Determine the number of offers over one week (equals two weeks here since there are only Oui and Non values in the availabilities_offers matrix)
N_o = sum(x in ["Oui", "Pair", "Impair"] for x in availabilities_offers.values.flatten())
print(f"Number of offers: {N_o}")

# Determine the number of requests over one week
N_r = sum(x in ["Oui", "Pair", "Impair"] for x in availabilities_requests.values.flatten())
print(f"Number of requests: {N_r}")



# Create the matrices O and R

# Initialize matrix O
O = np.zeros((N_o, 5), dtype=int)

# Counter for valid offers
idx = 0


# Go through the offers array

for i in range(availabilities_offers.shape[0]):
    nb_places = number_places_offers[i]
    for j in range(availabilities_offers.shape[1]):
        if availabilities_offers.iloc[i, j] == "Oui":
            # Store the offer in matrix O
            O[idx, 0] = i + 1
            O[idx, 1] = j + 1
            O[idx, 2] = 1
            O[idx, 3] = 1
            O[idx, 4] = nb_places
            idx += 1
        elif availabilities_offers.iloc[i, j] == "Pair":
            # Store the offer in matrix O
            O[idx, 0] = i + 1
            O[idx, 1] = j + 1
            O[idx, 2] = 1
            O[idx, 3] = 0
            O[idx, 4] = nb_places
            idx += 1
        elif availabilities_offers.iloc[i, j] == "Impair":
            # Store the offer in matrix O
            O[idx, 0] = i + 1
            O[idx, 1] = j + 1
            O[idx, 2] = 0
            O[idx, 3] = 1
            O[idx, 4] = nb_places
            idx += 1


print("Matrix O:")
print(O)
print()



# Initialize matrix R
R = np.zeros((N_r, 4), dtype=int)

# Counter for valid requests
idr = 0

# Go through the requests array
for i in range(availabilities_requests.shape[0]):
    for j in range(availabilities_requests.shape[1]):
        if availabilities_requests.iloc[i, j] == "Oui":
            # Store the request in matrix R
            R[idr, 0] = i + 1
            R[idr, 1] = j + 1
            R[idr, 2] = 1
            R[idr, 3] = 1
            idr += 1
        elif availabilities_requests.iloc[i, j] == "Pair":
            # Store the request in matrix R
            R[idr, 0] = i + 1
            R[idr, 1] = j + 1
            R[idr, 2] = 1
            R[idr, 3] = 0
            idr += 1
        elif availabilities_requests.iloc[i, j] == "Impair":
            # Store the request in matrix R
            R[idr, 0] = i + 1
            R[idr, 1] = j + 1
            R[idr, 2] = 0
            R[idr, 3] = 1
            idr += 1

print("Matrix R:")
print(R)
print()


# Display the available drivers and passengers at each hour

# Traverse all days and hours
for t in range(1, T+1):
    # Find available drivers at the hour and even weeks
    available_drivers_pairs = np.unique(O[(O[:, 1] == t) & (O[:, 2] == 1), 0])
    # Find available drivers at the hour and odd weeks
    available_drivers_impairs = np.unique(O[(O[:, 1] == t) & (O[:, 3] == 1), 0])

    # Display available drivers at the hour and even weeks
    print(f"Drivers available for a ride at hour {t} on even weeks:")
    print(available_drivers_pairs)
    
    # Display available drivers at the hour and odd weeks
    print(f"Drivers available for a ride at hour {t} on odd weeks:")
    print(available_drivers_impairs)

print()

for t in range(1, T+1):
    # Find available passengers at the hour and even weeks
    available_passengers_pairs = np.unique(R[(R[:, 1] == t) & (R[:, 2] == 1), 0])
    # Find available passengers at the hour and odd weeks
    available_passengers_impairs = np.unique(R[(R[:, 1] == t) & (R[:, 3] == 1), 0])

    # Display available passengers at the hour and even weeks
    print(f"Passengers requesting a ride at hour {t} on even weeks:")
    print(available_passengers_pairs)
    
    # Display available passengers at the hour and odd weeks
    print(f"Passengers requesting a ride at hour {t} on odd weeks:")
    print(available_passengers_impairs)



# A and B initialization with zeros
A = np.zeros((number_drivers, T, 2), dtype=int)
B = np.zeros((number_passengers, T, 2), dtype=int)


# Fill A with O values
for i in range(O.shape[0]):
    driver = O[i, 0] - 1  
    time = O[i, 1] - 1  
    if O[i, 2] == 1:
        A[driver, time, 0] = 1
    if O[i, 3] == 1:
        A[driver, time, 1] = 1


# Fill B with R values
for i in range(R.shape[0]):
    passenger = R[i, 0] - 1  
    time = R[i, 1] - 1  
    if R[i, 2] == 1:
        B[passenger, time, 0] = 1
    if R[i, 3] == 1:
        B[passenger, time, 1] = 1


print("Variable A:")
print(A)
print()
print("Variable B:")
print(B)



# Create the M matrix
M = np.zeros((number_passengers, number_drivers, T, 2), dtype=float)

Beta = 7

 
for child in range(number_passengers):
    for parent in range(number_drivers):
        if name_requests[child] in child_offers.iloc[parent].values: #### To check
            for t in range(T):
                for w in range(2):
                    M[child, parent, t, w] = Beta
        else:
            for t in range(T):
                for w in range(2):
                    M[child, parent, t, w] = 10 - Beta

print("Matrix M:")
with np.printoptions(threshold=np.inf):
    print(M)



############################################################ CARPOOLING FUNCTION ############################################################

def covoiturage(Beta, Alpha):
    
    # Create the statistics dictionary, in order to display it on a separate Excel sheet
    stat_df = {"Statistics": [], "Values": []}
    
    stat_df["Statistics"].append("Number of requests")
    stat_df["Values"].append(nb_requests)
    
    # Create the M matrix
    M = np.zeros((number_passengers, number_drivers, T, 2), dtype=float)
    
    for parent in range(name_offers.shape[0]):
        for child in range(name_requests.shape[0]):
            if name_requests[child] in child_offers.iloc[parent].values:
                for t in range(T):
                    for w in range(2):
                        M[child, parent, t, w] = Beta
            else:
                for t in range(T):
                    for w in range(2):
                        M[child, parent, t, w] = 10 - Beta
    

    # Create the optimization model
    m = LpProblem("Carpooling", LpMaximize)
    
    

    # Create the decision variables
    X = LpVariable.dicts("X", ((i, j, t, w) for i in range(0, number_passengers) for j in range(0, number_drivers) for t in range(0, T) for w in range(0, 2)), cat='Binary') 
    E = LpVariable.dicts("E", ((j, t, w) for j in range(0, number_drivers) for t in range(0, T) for w in range(0, 2)), cat='Binary')
    G = LpVariable.dicts("G", (j for j in range(0, number_drivers)), cat='Integer')

    # Objective function
    m += lpSum(X[i, j, t, w]*M[i, j, t, w] for i in range(0, number_passengers) for j in range(0, number_drivers) for t in range(0, T) for w in range(0, 2)) - lpSum(Alpha*G[j] for j in range(0, number_drivers)) 

    # Constraints
    for t in range(0, T):
        for w in range(0, 2):
            for c in range(0, number_drivers):
                m += E[c,t,w] >= (1/(number_passengers*T*2))*lpSum(X[n, c, t, w] for n in range(0, number_passengers))

    for c in range(0, number_drivers):
        m += lpSum(E[c,t, w] for t in range(0, T) for w in range(0, 2)) == G[c]

    for n in range(0, number_passengers):
        m += lpSum(X[n, c, t, w] for c in range(0, number_drivers) for t in range(0, T) for w in range(0, 2)) <= places_offered_passengers[n,1]

    for t in range(0, T):
        for c in range(0, number_drivers):
            for w in range(0, 2):
                m += lpSum(X[n, c, t, w] for n in range(0, number_passengers)) <= number_places_offers[c]

    for t in range(0, T):
        for n in range(0, number_passengers):
            for w in range(0, 2):
                m += lpSum(X[n, c, t, w]*A[c, t, w]*B[n, t, w] for c in range(0, number_drivers)) <= 1

    for t in range(0, T):
        for c in range(0, number_drivers):
            for w in range(0, 2):
                if A[c, t, w] == 0:
                    m += lpSum(X[n, c, t, w] for n in range(0, number_passengers)) == 0

    for t in range(0, T):
        for n in range(0, number_passengers):
            for w in range(0, 2):
                if B[n, t, w] == 0:
                    m += lpSum(X[n, c, t, w] for c in range(0, number_drivers)) == 0



    # Solve the optimization problem
    m.solve()

    # Print the status of the solution
    if m.status == 1:
        print("Solution optimale trouvée:")
    elif m.status == -1:
        print("Il n'existe aucune solution :'(")
    elif m.status == 0:
        print("Solution réalisable trouvée, mais pas nécessairement optimale")
    else:
        print("Voir la documentation")
        
        
    # Counting the number of satisfied requests
    nb_request_done = 0

    for w in range(0, 2):
        for t in range(0, T):
            for c in range(0, number_drivers):
                for n in range(0, number_passengers):
                    if X[n, c, t, w].varValue == 1:
                        nb_request_done += 1
       
    print("Value of the first variable X[0, 0, 0, 0]:")                 
    print(X[0, 0, 0, 0].varValue)
    
    print("Value of the first variable X[15, 4, 6, 1]:")
    print(X[15, 4, 6, 1].varValue)

    percentage_request = nb_request_done / nb_requests * 100
    
    stat_df["Statistics"].append("Number of requests satisfied")
    stat_df["Values"].append(nb_request_done)
    
    stat_df["Statistics"].append("Percentage of requests satisfied")
    stat_df["Values"].append(percentage_request)

    print()
    print("Le nombre de demande de conduite faite par les enfants est :", nb_requests)
    print("Le nombre de demande de conduite satisfaite est :", nb_request_done)
    print("Le pourcentage de demande satisfaite est :", percentage_request, "%")
    print()
        
    nb_place_dispo = 0
    for w in range(0, 2):
        for t in range(0, T):
            for c in range(0, number_drivers):
                res = 0
                for n in range(0, number_passengers):
                    if X[n, c, t, w].varValue == 1:
                        res += 1
                if res >= 1:
                    nb_place_dispo += number_places_offers[c]
                    
    percentage_remplissage = nb_request_done / nb_place_dispo * 100
    print(nb_place_dispo)
    
    stat_df["Statistics"].append("Percentage of car filling")
    stat_df["Values"].append(percentage_remplissage)

    print("Nombre de place disponible après répartition :", nb_place_dispo)
    print("Le taux de remplissage des voitures est :", percentage_remplissage, "%")
    print()
    
    I = 0
    L = 0
    E = 0
    for w in range(0, 2):
        for t in range(0, T):
            for c in range(0, number_drivers):
                J = 0
                for n in range(0, number_passengers):
                    L += X[n, c, t, w].varValue
                    J += X[n, c, t, w].varValue
                    if M[n, c, t, w] >= 5.01 and X[n, c, t, w].varValue == 1:
                        E += 1
                if J >= 1:
                    I += 1


    print("Nombre d'enfants sans leur parent :", E)
    print()

    pourcent_family = (1 - (E / L)) * 100

    stat_df["Statistics"].append("Percentage of children in their parent's car")
    stat_df["Values"].append(pourcent_family)

    print("Pourcentage d'enfant dans la voiture de leur parent :", pourcent_family, "%")
    print()

    print("Nombre de voitures utilisée :", I, ", Nombre d'enfants dans les voitures :", L, ", Nombre d'enfants sans leur parent :", E)
    print()
    
    ############################## Display the results ######################################
    ############################## Create a result Excel table ##############################
    
    
    # Create a table to display the results
    W = [["" for _ in range(T)] for _ in range(300)]

    # Add the schedules
    for i in range(T):
        W[0][i] = horaires_requests[i]

    # Go through all the schedules (Monday 8AM, ...)
    for t in range(T):
        h = horaires_requests[t]
        Time = f"{h}"
        print()
        print(Time)
        VAR1 = []
        
        # Go through the drivers
        for j in range(number_drivers):
            nb = number_places_offers[j]
            d = name_offers[j]
            e = number_places_offers[j]
            f = f"{d} {e} places"
            offer = ""
            
            # Go through the weeks
            for w in range(0, 2):
                VAR2 = []
                VAR3 = []
                
                if w == 0:
                    week = "Semaine paire"
                elif w == 1:
                    week = "Semaine Impaire"
                
                VAR2.append(week)
                VAR2.append(f)
                VAR3.append(week)
                VAR3.append(f)
                offer = f"{d} with {nb} places, week {week}:"
                
                # Go through the requests to find the passengers
                for i in range(number_passengers):
                    n = name_requests[i]
                    if X[i, j, t, w].varValue == 1:
                        offer += f"{n}, "
                        VAR2.append(n)
                
                if offer != f"{d} with {nb} places, week {week}:": 
                    print(offer)
                
                if VAR2 != VAR3:
                    length = len(VAR2)
                    VAR1.extend(VAR2)
                    Q = nb + 2 - length  # Determine the number of empty spots in the car
                    for _ in range(Q):
                        VAR1.append("*")  # "*" for all empty spots in the car
                            
            if VAR1:
                for g in range(len(VAR1)):
                    W[g+1][t] = VAR1[g]

    print()

    # Display the number of trips per child
    nombre_trajets_par_enfant = [0] * number_passengers

    # Traverse the indices n, c, t, w
    for n in range(number_passengers):
        for c in range(number_drivers):
            for t in range(T):
                for w in range(0, 2):
                    if X[n, c, t, w].varValue == 1:
                        nombre_trajets_par_enfant[n] += 1

    for n in range(number_passengers):
        nbu = places_offered_passengers[n][1]
        print(f"Enfant {n} qui a le droit à {nbu} trajets: {nombre_trajets_par_enfant[n]} trajets")

    #print(places_offered_passengers)

    #### Display both the driving schedule and the statistics ####
    df = pd.DataFrame(W)
    stat_df = pd.DataFrame(stat_df)
    with pd.ExcelWriter(f'./Repartition_Voiture_{year}.xlsx') as writer:
        df.to_excel(writer, sheet_name='Planning', index=False, header=False)
        stat_df.to_excel(writer, sheet_name='Statistics', index=False, header=True)
        
        



# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook(f'./Repartition_Voiture_{year}.xlsx')
schedule_sheet = workbook.active

        
        
def format_schedule():
    week = ["lundi", "mardi", "mercredi", "jeudi", "vendredi"]

    # Loop through all the cells in the sheet
    for row in schedule_sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                
                for i in week:
                    if i in cell.value.lower():
                        cell.fill = dark_grey_fill
                
                if 'semaine paire' in cell.value.lower() :
                    cell.fill = dark_blue_fill
                elif 'semaine impaire' in cell.value.lower():
                    cell.fill = light_blue_fill
                elif 'places' in cell.value.lower():
                    cell.fill = red_fill
                elif '*' in cell.value.lower():
                    cell.fill = light_yellow_fill

    for column in schedule_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.0
        schedule_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    statistics_sheet = workbook["Statistics"]
    for column in statistics_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.0
        statistics_sheet.column_dimensions[column[0].column_letter].width = adjusted_width


    workbook.save(f'./Repartition_Voiture_{year}_vf.xlsx')
        

covoiturage(6.5, 4)
format_schedule()
print()
print("Excel sheet created successfully")